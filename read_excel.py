import sys
import codecs

try:
    from openpyxl import load_workbook
    wb = load_workbook(r'C:\Users\gw\Downloads\雷池-3000 Datasheet.xlsx', data_only=True)
    ws = wb.active
    print('Sheet name:', ws.title)
    print('\nAll data:')
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        # Filter out None values
        row_values = [str(cell) if cell is not None else '' for cell in row]
        if any('throughput' in str(v).lower() or '吞吐' in str(v) or 'bps' in str(v).lower() or 'Gbps' in str(v) for v in row_values):
            print(f'Row {i}:', row_values)
        # Print all rows to see structure
        print(f'Row {i}:', row_values)
except Exception as e:
    print(f'Error: {e}')
    import traceback
    traceback.print_exc()
